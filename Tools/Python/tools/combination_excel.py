import warnings

import openpyxl
from numpy.f2py.crackfortran import endifs
from soupsieve.util import lower

warnings.filterwarnings("ignore", category=UserWarning)

import pandas as pd
import sys
import os

# 读取Excel文件
# 当前项目文件路径
current_dir = sys.path[0]
# 配置文件存放的文件夹
config_folder = os.path.join(current_dir, "Config")

# 输出文件存放的文件目录
export_folder = os.path.join(current_dir, 'Output')

# 创建文件夹
def create_folder(folder_name):
    """
    创建文件夹
    :param folder_name: 文件夹名
    :return:
    """
    folder_path = os.path.join(export_folder, folder_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return folder_path


def output():

    ddl_folder = create_folder(os.path.join(config_folder, 'model_design'))
    src_excel_file = os.path.join(ddl_folder, '架构详细设计-DIM.xlsx') # 源Excel

    # 输出文件存放的文件目录
    export_folder = create_folder(os.path.join(current_dir, 'Output', 'combined'))
    tgt_excel_file = os.path.join(export_folder, 'combined_model_design.xlsx') # 目标Excel

    wb = openpyxl.load_workbook(src_excel_file)

    # 指定开始的工作表名称
    start_sheet_name = '模板'  # 请替换为你的开始工作表名称

    # 指定行列范围
    start_row = 9  # 从第8行开始（0-based索引）
    start_col = 3  # 从第3列开始（0-based索引）
    end_col = 5  # 到第10列结束（0-based索引）

    # 获取所有工作表名称
    sheet_names = wb.sheetnames

    # 找到开始工作表的索引
    start_index = sheet_names.index(start_sheet_name) + 1

    # 创建一个新的工作表来存储整合后的数据
    combined_sheet = wb.create_sheet(title='Combined')

    # 遍历从指定工作表开始的所有工作表并将其内容添加到combined_sheet中
    for sheet_name in sheet_names[start_index:]:
        sheet = wb[sheet_name]

        schema_name =  lower(sheet[3][1].value[:3])
        table_name = sheet[3][1].value
        table_desc = sheet[4][1].value

        if schema_name == 'dim':
            schema_name = 'kwi_dwr'
        else :
            schema_name = 'kwi_' + schema_name

        # print(sheet_name, ":", schema_name, table_name, table_desc)

        combined_sheet.append(schema_name)
        combined_sheet.append(table_name)
        combined_sheet.append(table_desc)

        # 选择指定范围内的数据
        for row in sheet.iter_rows(min_row=start_row, min_col=start_col, max_col=end_col):
            row_data = [cell.value for cell in row]


            combined_sheet.append(row_data)

    # 保存新的Excel文件
    wb.save(tgt_excel_file)

    # print(f"数据已整合并保存到 {tgt_excel_file}")


def output1():


    # 读取Excel文件
    ddl_folder = create_folder(os.path.join(config_folder, 'model_design'))
    file_path = os.path.join(ddl_folder, '架构详细设计-DIM.xlsx') # 源Excel

    # 输出文件存放的文件目录
    export_folder = create_folder(os.path.join(current_dir, 'Output', 'combined'))
    new_file_path = os.path.join(export_folder, 'combined_model_design.xlsx') # 目标Excel


    # 读取Excel文件
    sheet_name = 'DWI_产品发货量折算系数配置表'  # 替换为你的工作表名称

    # 读取B2:B4单元格的值
    df_b = pd.read_excel(file_path, sheet_name=sheet_name, usecols='B', skiprows=1, nrows=3, header=None)

    # 读取C:E列的值
    df_ce = pd.read_excel(file_path, sheet_name=sheet_name, usecols='C:E', skiprows=7, header=None)

    # 重置索引
    df_b.reset_index(drop=True, inplace=True)
    df_ce.reset_index(drop=True, inplace=True)

    # 拼接B2:B4和C:E列
    df_new = pd.concat([df_b, df_ce], axis=1)

    # 保存新表格到新的Excel文件
    df_new.to_excel(new_file_path, index=False, header=['B', 'C', 'D', 'E'])

    print("新表格已保存到:", new_file_path)

if __name__ == '__main__':
    output1()