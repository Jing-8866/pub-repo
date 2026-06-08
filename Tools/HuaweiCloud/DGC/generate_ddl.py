#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""根据架构详细设计 Excel 生成华为云 DWS 建表 DDL。"""

import glob
import os
from datetime import datetime

import openpyxl

PATH = os.path.dirname(os.path.abspath(__file__)) # excel 文件所在目录 C:\Self\Personally\Script\AI\KLYY
OUT_DIR = os.path.join(PATH, "DDL") # DDL 文件输出目录 C:\Self\Personally\Script\AI\KLYY\DDL
WITH_CLAUSE = ( # DWS 建表 DDL 的 WITH 子句
    "WITH (ORIENTATION = column, COMPRESSION = MIDDLE, "
    "colversion=3.0, enable_hstore_opt = true) DISTRIBUTE BY Roundrobin"
) # 存储方式、压缩方式、分布方式


def escape_comment(text: str) -> str: # 转义注释中的单引号
    return text.replace("'", "''") if text else ""


def resolve_duplicate_name(name: str, comment: str, seen: set[str]) -> tuple[str, str]: # 解决字段名重复的问题
    col = name.lower()
    if col not in seen:
        seen.add(col)
        return col, comment

    if "预计" in comment or "预测" in comment:
        col = f"{col}_forecast"
    else:
        col = f"{col}_hist"

    seen.add(col)
    note = f"{comment}（设计文档字段名重复，已重命名为{col}）"
    return col, note


def generate_ddl(table_full: str, table_desc: str, fields: list[tuple[str, str, str]]) -> str: # 生成 DWS 建表 DDL
    schema, table = table_full.lower().split(".", 1) # 表名拆分，schema 为 schema 名称，table 为表名
    lines = [
        f"""-- DWS sql 
-- ******************************************************************** --
-- author: 
-- create time: {datetime.now().strftime("%Y/%m/%d %H:%M:%S")} GMT+08:00
-- {table_desc}
-- ******************************************************************** --
        """,
        "",
        f"DROP TABLE IF EXISTS {schema}.{table};",
        "",
        f"CREATE TABLE {schema}.{table} (",
    ]

    seen: set[str] = set() # 字段名集合
    col_defs = []
    for name, dtype, comment in fields: # 遍历字段列表
        col, cmt = resolve_duplicate_name(name, comment, seen) # 解决字段名重复的问题
        field = f"    {col} {dtype.lower()}" # 字段定义 
        if cmt: # 如果字段注释不为空，则添加注释
            field += f" comment '{escape_comment(cmt)}'"
        col_defs.append(field)

    lines.append(",\n".join(col_defs)) # 将字段定义拼接成字符串
    lines.append(")")
    lines.append(WITH_CLAUSE) # 添加存储方式、压缩方式、分布方式
    lines.append(f"comment '{escape_comment(table_desc)}'") # 添加表注释
    lines.append(";") # 添加分号
    return "\n".join(lines) # 将 DDL 拼接成字符串


def main() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)  # 创建输出目录
    xlsx_files = [
        f for f in glob.glob(os.path.join(PATH, "*.xlsx"))
        if not os.path.basename(f).startswith("~")
    ]
    if not xlsx_files:
        raise FileNotFoundError("未找到架构详细设计 Excel 文件")

    # 获取用户输入
    input_text = input("请输入要生成的 Excel 文件名（多个用逗号分隔，留空生成所有）: ").strip()
    
    if input_text:
        # 分割并清理输入
        search_terms = [term.strip() for term in input_text.split(',') if term.strip()]
        
        # 查找匹配的文件
        file_list = []
        unmatched_terms = []
        
        for term in search_terms:
            # 查找包含搜索词的文件
            matched = [f for f in xlsx_files if term in os.path.basename(f)]
            if matched:
                file_list.extend(matched)
            else:
                unmatched_terms.append(term)
        
        # 处理未匹配的搜索词
        if unmatched_terms:
            print(f"\n警告: 未找到以下文件: {', '.join(unmatched_terms)}")
        
        if not file_list:
            print("未匹配到任何文件")
            # file_list = xlsx_files
        else:
            # 去重
            file_list = list(dict.fromkeys(file_list))
    else:
        print("未输入文件名，生成所有 Excel 文件")
        file_list = xlsx_files
    
    # 显示将要处理的文件
    print(f"\n将处理 {len(file_list)} 个文件:")
    for i, file_path in enumerate(file_list, 1):
        print(f"{i}. {os.path.basename(file_path)}")
    
    # 遍历处理所有文件
    for file_path in file_list:
        file_name = os.path.basename(file_path)
        print(f"\n正在处理: {file_path}")
        generate_ddl_file(file_path) # 生成 DDL
        
        try:
            # 这里添加处理文件的代码
            # process_excel_file(file_path)
            pass
        except Exception as e:
            print(f"处理文件 {file_name} 时出错: {e}")



def generate_ddl_file(file_path: str) -> str: # 生成 DDL
    wb = openpyxl.load_workbook(file_path, data_only=True) # 加载 Excel 文件
    all_sql: list[str] = [] # 所有 DDL 列表

    for sheet_name in wb.sheetnames: # 遍历所有 sheet 页
        if sheet_name == "目录": # 排除目录页
            continue

        ws = wb[sheet_name] # 获取 sheet 页
        table_full = str(ws.cell(3, 2).value).strip() # 获取表名
        table_desc = str(ws.cell(4, 2).value).strip() # 获取表描述
        fields: list[tuple[str, str, str]] = [] # 字段列表

        for row in range(9, ws.max_row + 1): # 遍历所有行
            fname = ws.cell(row, 3).value # 获取字段名
            ftype = ws.cell(row, 4).value # 获取字段类型
            fcomment = ws.cell(row, 5).value # 获取字段注释
            if fname and ftype: # 如果字段名和字段类型不为空，则添加到字段列表
                fields.append((
                    str(fname).strip(), # 字段名
                    str(ftype).strip(), # 字段类型 去除空格
                    str(fcomment).strip() if fcomment else "", # 字段注释
                ))

        ddl = generate_ddl(table_full, table_desc, fields) # 生成 DDL
        table_name = table_full.split(".")[-1].lower() # 获取表名 小写
        out_file = os.path.join(OUT_DIR, f"{table_name}.sql") # 输出文件路径
        with open(out_file, "w", encoding="utf-8") as f: # 写入文件
            f.write(ddl + "\n") # 写入 DDL
        all_sql.append(ddl) # 添加 DDL 到列表
        print(f"generated: {out_file}") # 打印生成文件路径

    # combined = os.path.join(OUT_DIR, "all_tables.sql") # 所有 DDL 文件路径
    # with open(combined, "w", encoding="utf-8") as f: # 写入文件
    #     f.write("\n\n".join(all_sql) + "\n") # 写入所有 DDL
    # print(f"generated: {combined}") # 打印生成文件路径


if __name__ == "__main__":
    main()
