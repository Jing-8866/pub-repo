#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""根据架构详细设计 Excel 生成华为云 DWS 建表 DDL。"""

import glob
import os
import sys
import tkinter as tk
from datetime import datetime
from tkinter import messagebox

import openpyxl

WITH_CLAUSE = (
    "WITH (ORIENTATION = column, COMPRESSION = MIDDLE, "
    "colversion=3.0, enable_hstore_opt = true) DISTRIBUTE BY Roundrobin"
)

TABLE_NAME_LABELS = ("表名称", "表名")
TABLE_DESC_LABELS = ("表描述",)
FIELD_HEADER_LABELS = ("字段名称", "字段名")
TABLE_TYPE_LABELS = ("表类型",)


def get_app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


PATH = get_app_dir()
OUT_DIR = os.path.join(PATH, "DDL")


def escape_comment(text: str) -> str:
    return text.replace("'", "''") if text else ""


def generate_script_header(desc: str) -> str:
    return f"""-- DWS sql 
-- ******************************************************************** --
-- author: 
-- create time: {datetime.now().strftime("%Y/%m/%d %H:%M:%S")} GMT+08:00
-- {desc}
-- ******************************************************************** --
"""


def find_duplicate_fields(fields: list[tuple[str, str, str]]) -> list[str]:
    seen: set[str] = set()
    duplicates: list[str] = []
    for name, _, _ in fields:
        col = name.lower()
        if col in seen and col not in duplicates:
            duplicates.append(col)
        seen.add(col)
    return duplicates


def generate_ddl(table_full: str, table_desc: str, fields: list[tuple[str, str, str]]) -> str:
    schema, table = table_full.lower().split(".", 1)
    lines = [
        f"DROP TABLE IF EXISTS {schema}.{table};",
        "",
        f"CREATE TABLE {schema}.{table} (",
    ]

    col_defs = []
    for name, dtype, comment in fields:
        col = name.lower()
        field = f"    {col} {dtype.lower()}"
        if comment:
            field += f" comment '{escape_comment(comment)}'"
        col_defs.append(field)

    lines.append(",\n".join(col_defs))
    lines.append(")")
    lines.append(WITH_CLAUSE)
    lines.append(f"comment '{escape_comment(table_desc)}'")
    lines.append(";")
    return "\n".join(lines)


def _cell_text(ws, row: int, col: int) -> str:
    value = ws.cell(row, col).value
    if value is None:
        return ""
    return str(value).strip()


def _is_table_name_row(ws, row: int) -> bool:
    label = _cell_text(ws, row, 1)
    table_full = _cell_text(ws, row, 2)
    return label in TABLE_NAME_LABELS and "." in table_full


def _is_table_block_boundary(ws, row: int) -> bool:
    label = _cell_text(ws, row, 1)
    if label in TABLE_TYPE_LABELS:
        return True
    return _is_table_name_row(ws, row)


def _find_field_header_row(ws, start_row: int, end_row: int) -> int | None:
    for row in range(start_row, end_row + 1):
        header = _cell_text(ws, row, 3)
        if header in FIELD_HEADER_LABELS:
            return row
    return None


def _is_empty_field_row(ws, row: int) -> bool:
    return not any(_cell_text(ws, row, col) for col in (3, 4, 5))


def parse_sheet_tables(ws) -> list[tuple[str, str, list[tuple[str, str, str]]]]:
    table_name_rows = [
        row for row in range(1, ws.max_row + 1) if _is_table_name_row(ws, row)
    ]
    if not table_name_rows:
        return []

    tables: list[tuple[str, str, list[tuple[str, str, str]]]] = []
    for idx, name_row in enumerate(table_name_rows):
        next_name_row = (
            table_name_rows[idx + 1] if idx + 1 < len(table_name_rows) else ws.max_row + 1
        )

        table_full = _cell_text(ws, name_row, 2)
        table_desc = ""
        search_start = name_row + 1
        if search_start < next_name_row and _cell_text(ws, search_start, 1) in TABLE_DESC_LABELS:
            table_desc = _cell_text(ws, search_start, 2)
            search_start += 1

        header_row = _find_field_header_row(ws, search_start, next_name_row - 1)
        if header_row is None:
            continue

        fields: list[tuple[str, str, str]] = []
        for row in range(header_row + 1, next_name_row):
            if _is_table_block_boundary(ws, row):
                break
            if _is_empty_field_row(ws, row):
                break

            fname = ws.cell(row, 3).value
            ftype = ws.cell(row, 4).value
            fcomment = ws.cell(row, 5).value
            if fname and ftype:
                fields.append((
                    str(fname).strip(),
                    str(ftype).strip(),
                    str(fcomment).strip() if fcomment else "",
                ))

        if fields:
            tables.append((table_full, table_desc, fields))

    return tables


def generate_ddl_file(file_path: str) -> dict:
    result = {"paths": [], "errors": [], "script_count": 0, "table_count": 0}
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name == "目录":
            continue

        ws = wb[sheet_name]
        tables = parse_sheet_tables(ws)
        if not tables:
            result["errors"].append(f"sheet '{sheet_name}' 无有效表结构")
            continue

        seen_tables: set[str] = set()
        ddls: list[str] = []

        for table_full, table_desc, fields in tables:
            table_key = table_full.lower()
            if table_key in seen_tables:
                result["errors"].append(
                    f"表名重复，跳过: {table_full} (sheet: {sheet_name})"
                )
                continue
            seen_tables.add(table_key)

            dup_fields = find_duplicate_fields(fields)
            if dup_fields:
                result["errors"].append(
                    f"字段名重复 {', '.join(dup_fields)}，跳过: "
                    f"{table_full} (sheet: {sheet_name})"
                )
                continue

            ddls.append(generate_ddl(table_full, table_desc, fields))

        if not ddls:
            result["errors"].append(f"sheet '{sheet_name}' 无有效表结构可生成")
            continue

        first_table_name = tables[0][0].split(".")[-1].lower()
        out_file = os.path.join(OUT_DIR, f"{first_table_name}.sql")
        script_parts = [generate_script_header(tables[0][1]), *ddls]

        with open(out_file, "w", encoding="utf-8") as f:
            f.write("\n\n".join(script_parts) + "\n")

        result["paths"].append(out_file)
        result["script_count"] += 1
        result["table_count"] += len(ddls)

    return result


def format_summary(paths: list[str], errors: list[str], script_count: int, table_count: int) -> str:
    lines = [f"执行时间: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}", ""]

    if paths:
        out_dirs = list(dict.fromkeys(os.path.dirname(p) for p in paths))
        lines.append("输出路径:")
        for out_dir in out_dirs:
            lines.append(f"  {out_dir}")
        lines.append("")

    if errors:
        lines.append("错误信息:")
        for error in errors:
            lines.append(f"  {error}")
        lines.append("")

    lines.append(f"脚本数量: {script_count}")
    lines.append(f"表数量: {table_count}")
    return "\n".join(lines)


def list_xlsx_files() -> list[str]:
    return sorted(
        f for f in glob.glob(os.path.join(PATH, "*.xlsx"))
        if not os.path.basename(f).startswith("~")
    )


def select_xlsx_files(xlsx_files: list[str]) -> list[str] | None:
    selected: list[str] = []
    root = tk.Tk()
    root.title("DDL 生成工具 - 选择 Excel 文件")
    root.geometry("560x420")
    root.resizable(True, True)

    tk.Label(
        root,
        text="请选择要生成的 Excel 文件（按住 Ctrl 可多选）:",
        anchor="w",
    ).pack(fill=tk.X, padx=12, pady=(12, 6))

    frame = tk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True, padx=12)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    listbox = tk.Listbox(
        frame,
        selectmode=tk.EXTENDED,
        yscrollcommand=scrollbar.set,
        font=("Microsoft YaHei UI", 10),
    )
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=listbox.yview)

    for file_path in xlsx_files:
        listbox.insert(tk.END, os.path.basename(file_path))

    if xlsx_files:
        listbox.selection_set(0, tk.END)

    result: list[str | None] = [None]

    def on_confirm() -> None:
        indices = listbox.curselection()
        if not indices:
            messagebox.showwarning("提示", "请至少选择一个 Excel 文件", parent=root)
            return
        for index in indices:
            selected.append(xlsx_files[index])
        result[0] = selected
        root.destroy()

    def on_cancel() -> None:
        root.destroy()

    btn_frame = tk.Frame(root)
    btn_frame.pack(pady=12)
    tk.Button(btn_frame, text="生成", width=12, command=on_confirm).pack(side=tk.LEFT, padx=6)
    tk.Button(btn_frame, text="取消", width=12, command=on_cancel).pack(side=tk.LEFT, padx=6)

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    root.mainloop()
    return result[0]


def write_log(content: str) -> str:
    log_path = os.path.join(PATH, f"generate_ddl_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write(content + "\n")
    return log_path


def run() -> str:
    os.makedirs(OUT_DIR, exist_ok=True)
    paths: list[str] = []
    errors: list[str] = []
    script_count = 0
    table_count = 0

    xlsx_files = list_xlsx_files()
    if not xlsx_files:
        summary = format_summary([], ["未找到同目录下的 Excel 文件（*.xlsx）"], 0, 0)
        return write_log(summary)

    file_list = select_xlsx_files(xlsx_files)
    if file_list is None:
        summary = format_summary([], ["用户取消操作"], 0, 0)
        return write_log(summary)

    for file_path in file_list:
        file_name = os.path.basename(file_path)
        try:
            result = generate_ddl_file(file_path)
            paths.extend(result["paths"])
            errors.extend(result["errors"])
            script_count += result["script_count"]
            table_count += result["table_count"]
        except Exception as e:
            errors.append(f"处理文件 {file_name} 时出错: {e}")

    summary = format_summary(paths, errors, script_count, table_count)
    return write_log(summary)


def main() -> None:
    log_path = run()
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("完成", f"执行完成，日志已保存至:\n{log_path}")
    root.destroy()


if __name__ == "__main__":
    main()
