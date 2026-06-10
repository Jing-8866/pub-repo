#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""根据架构详细设计 Excel 生成华为云 DWS 建表 DDL。"""

import glob
import os
from datetime import datetime

import openpyxl

PATH = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(PATH, "DDL")
WITH_CLAUSE = (
    "WITH (ORIENTATION = column, COMPRESSION = MIDDLE, "
    "colversion=3.0, enable_hstore_opt = true) DISTRIBUTE BY Roundrobin"
)

TABLE_NAME_LABELS = ("表名称", "表名")
TABLE_DESC_LABELS = ("表描述",)
FIELD_HEADER_LABELS = ("字段名称", "字段名")
TABLE_TYPE_LABELS = ("表类型",)


def escape_comment(text: str) -> str:
    return text.replace("'", "''") if text else ""


def generate_script_header(desc: str) -> str:
    return f"""-- DWS sql 
-- ******************************************************************** --
-- author: 
-- create time: {datetime.now().strftime("%Y/%m/%d %H:%M:%S")} GMT+08:00
-- {desc}
-- ******************************************************************** --
"""


def find_duplicate_fields(fields: list[tuple[str, str, str]]) -> list[str]:
    seen: set[str] = set()
    duplicates: list[str] = []
    for name, _, _ in fields:
        col = name.lower()
        if col in seen and col not in duplicates:
            duplicates.append(col)
        seen.add(col)
    return duplicates


def generate_ddl(table_full: str, table_desc: str, fields: list[tuple[str, str, str]]) -> str:
    schema, table = table_full.lower().split(".", 1)
    lines = [
        f"DROP TABLE IF EXISTS {schema}.{table};",
        "",
        f"CREATE TABLE {schema}.{table} (",
    ]

    col_defs = []
    for name, dtype, comment in fields:
        col = name.lower()
        field = f"    {col} {dtype.lower()}"
        if comment:
            field += f" comment '{escape_comment(comment)}'"
        col_defs.append(field)

    lines.append(",\n".join(col_defs))
    lines.append(")")
    lines.append(WITH_CLAUSE)
    lines.append(f"comment '{escape_comment(table_desc)}'")
    lines.append(";")
    return "\n".join(lines)


def _cell_text(ws, row: int, col: int) -> str:
    value = ws.cell(row, col).value
    if value is None:
        return ""
    return str(value).strip()


def _is_table_name_row(ws, row: int) -> bool:
    label = _cell_text(ws, row, 1)
    table_full = _cell_text(ws, row, 2)
    return label in TABLE_NAME_LABELS and "." in table_full


def _is_table_block_boundary(ws, row: int) -> bool:
    label = _cell_text(ws, row, 1)
    if label in TABLE_TYPE_LABELS:
        return True
    return _is_table_name_row(ws, row)


def _find_field_header_row(ws, start_row: int, end_row: int) -> int | None:
    for row in range(start_row, end_row + 1):
        header = _cell_text(ws, row, 3)
        if header in FIELD_HEADER_LABELS:
            return row
    return None


def _is_empty_field_row(ws, row: int) -> bool:
    """字段区空行：C/D/E 列均无有效内容。"""
    return not any(_cell_text(ws, row, col) for col in (3, 4, 5))


def parse_sheet_tables(ws) -> list[tuple[str, str, list[tuple[str, str, str]]]]:
    """解析单个 sheet 页中的全部表结构。"""
    table_name_rows = [
        row for row in range(1, ws.max_row + 1) if _is_table_name_row(ws, row)
    ]
    if not table_name_rows:
        return []

    tables: list[tuple[str, str, list[tuple[str, str, str]]]] = []
    for idx, name_row in enumerate(table_name_rows):
        next_name_row = (
            table_name_rows[idx + 1] if idx + 1 < len(table_name_rows) else ws.max_row + 1
        )

        table_full = _cell_text(ws, name_row, 2)
        table_desc = ""
        search_start = name_row + 1
        if search_start < next_name_row and _cell_text(ws, search_start, 1) in TABLE_DESC_LABELS:
            table_desc = _cell_text(ws, search_start, 2)
            search_start += 1

        header_row = _find_field_header_row(ws, search_start, next_name_row - 1)
        if header_row is None:
            continue

        fields: list[tuple[str, str, str]] = []
        for row in range(header_row + 1, next_name_row):
            if _is_table_block_boundary(ws, row):
                break
            if _is_empty_field_row(ws, row):
                break

            fname = ws.cell(row, 3).value
            ftype = ws.cell(row, 4).value
            fcomment = ws.cell(row, 5).value
            if fname and ftype:
                fields.append((
                    str(fname).strip(),
                    str(ftype).strip(),
                    str(fcomment).strip() if fcomment else "",
                ))

        if fields:
            tables.append((table_full, table_desc, fields))

    return tables


def generate_ddl_file(file_path: str) -> dict:
    """生成 DDL 文件，返回路径、错误及成功统计。"""
    result = {"paths": [], "errors": [], "script_count": 0, "table_count": 0}
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name == "目录":
            continue

        ws = wb[sheet_name]
        tables = parse_sheet_tables(ws)
        if not tables:
            result["errors"].append(f"sheet '{sheet_name}' 无有效表结构")
            continue

        seen_tables: set[str] = set()
        ddls: list[str] = []

        for table_full, table_desc, fields in tables:
            table_key = table_full.lower()
            if table_key in seen_tables:
                result["errors"].append(
                    f"表名重复，跳过: {table_full} (sheet: {sheet_name})"
                )
                continue
            seen_tables.add(table_key)

            dup_fields = find_duplicate_fields(fields)
            if dup_fields:
                result["errors"].append(
                    f"字段名重复 {', '.join(dup_fields)}，跳过: "
                    f"{table_full} (sheet: {sheet_name})"
                )
                continue

            ddls.append(generate_ddl(table_full, table_desc, fields))

        if not ddls:
            result["errors"].append(f"sheet '{sheet_name}' 无有效表结构可生成")
            continue

        first_table_name = tables[0][0].split(".")[-1].lower()
        out_file = os.path.join(OUT_DIR, f"{first_table_name}.sql")
        script_parts = [generate_script_header(tables[0][1]), *ddls]

        with open(out_file, "w", encoding="utf-8") as f:
            f.write("\n\n".join(script_parts) + "\n")

        result["paths"].append(out_file)
        result["script_count"] += 1
        result["table_count"] += len(ddls)

    return result


def print_summary(paths: list[str], errors: list[str], script_count: int, table_count: int) -> None:
    if paths:
        out_dirs = list(dict.fromkeys(os.path.dirname(p) for p in paths))
        print("输出路径:")
        for out_dir in out_dirs:
            print(f"  {out_dir}")

    if errors:
        print("错误信息:")
        for error in errors:
            print(f"  {error}")

    print(f"脚本数量: {script_count}")
    print(f"表数量: {table_count}")


def main() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)
    paths: list[str] = []
    errors: list[str] = []
    script_count = 0
    table_count = 0

    xlsx_files = [
        f for f in glob.glob(os.path.join(PATH, "*.xlsx"))
        if not os.path.basename(f).startswith("~")
    ]
    if not xlsx_files:
        print_summary([], ["未找到架构详细设计 Excel 文件"], 0, 0)
        return

    input_text = input("请输入要生成的 Excel 文件名（多个用逗号分隔，留空生成所有）: ").strip()

    if input_text:
        search_terms = [term.strip() for term in input_text.split(",") if term.strip()]
        file_list = []
        unmatched_terms = []

        for term in search_terms:
            matched = [f for f in xlsx_files if term in os.path.basename(f)]
            if matched:
                file_list.extend(matched)
            else:
                unmatched_terms.append(term)

        if unmatched_terms:
            errors.append(f"未找到以下文件: {', '.join(unmatched_terms)}")

        if not file_list:
            errors.append("未匹配到任何文件")
            file_list = []
        else:
            file_list = list(dict.fromkeys(file_list))
    else:
        file_list = xlsx_files

    for file_path in file_list:
        file_name = os.path.basename(file_path)
        try:
            result = generate_ddl_file(file_path)
            paths.extend(result["paths"])
            errors.extend(result["errors"])
            script_count += result["script_count"]
            table_count += result["table_count"]
        except Exception as e:
            errors.append(f"处理文件 {file_name} 时出错: {e}")

    print_summary(paths, errors, script_count, table_count)


if __name__ == "__main__":
    main()
